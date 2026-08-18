import os
import glob
import openpyxl
import pandas as pd
from typing import List, Dict, Any, Tuple, Optional
from app.core.config import settings

class ExcelConsolidationEngine:
    def __init__(self, source_dir: Optional[str] = None, output_dir: Optional[str] = None):
        self.source_dir = source_dir or settings.UNPROTECTED_DIR
        self.output_dir = output_dir or settings.OUTPUT_DIR
        os.makedirs(self.output_dir, exist_ok=True)

    def get_file_list(self) -> List[str]:
        pattern = os.path.join(self.source_dir, "*.[xX][lL][sS]*")
        return glob.glob(pattern)

    def inspect_files_and_sheets(self) -> Dict[str, Any]:
        files = self.get_file_list()
        file_details = []
        all_unique_sheets = set()

        for file_path in files:
            filename = os.path.basename(file_path)
            try:
                xls = pd.ExcelFile(file_path)
                sheet_names = xls.sheet_names
                all_unique_sheets.update(sheet_names)
                file_details.append({"filename": filename, "sheet_names": sheet_names})
            except Exception as e:
                file_details.append({"filename": filename, "error": str(e), "sheet_names": []})

        default_sheet = "سهام" if "سهام" in all_unique_sheets else (list(all_unique_sheets)[0] if all_unique_sheets else "سهام")
        return {"total_files": len(files), "available_sheets": sorted(list(all_unique_sheets)), "default_selected_sheet": default_sheet}

    def consolidate(self, target_sheets: List[str] = ["سهام"], output_filename: str = "consolidated_portfolio.xlsx") -> Dict[str, Any]:
        files = self.get_file_list()
        if not files:
            raise ValueError(f"هیچ فایل اکسلی در مسیر {self.source_dir} یافت نشد.")

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)

        sheet1_raw = out_wb.create_sheet(title="تلفیق خام (با سربرگ)")
        sheet2_clean = out_wb.create_sheet(title="تلفیق تمیز (بدون سربرگ)")

        raw_row_cursor = 1
        clean_row_cursor = 1
        clean_header_written = False
        total_processed = 0

        for file_path in files:
            filename = os.path.basename(file_path)
            fund_name = filename.replace("_latest", "").replace("_portfolio", "").replace(".xlsx", "").replace("_", " ")

            try:
                src_wb = openpyxl.load_workbook(file_path, data_only=True)
            except Exception:
                continue

            target_ws = None
            for sheet_name in target_sheets:
                if sheet_name in src_wb.sheetnames:
                    target_ws = src_wb[sheet_name]
                    break

            if not target_ws and len(src_wb.sheetnames) > 0:
                target_ws = src_wb.worksheets[0]

            if not target_ws:
                continue

            total_processed += 1

            # Step 3: Raw Concat
            for row in target_ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    for col_idx, cell_value in enumerate(row, 1):
                        sheet1_raw.cell(row=raw_row_cursor, column=col_idx, value=cell_value)
                    raw_row_cursor += 1
            raw_row_cursor += 1

            # Step 4: Clean Concat
            grid = [list(row) for row in target_ws.iter_rows(values_only=True)]
            header_idx, data_start_idx = self._detect_table_bounds(grid)

            if header_idx is not None and not clean_header_written:
                header_row = ["نام صندوق"] + grid[header_idx]
                for col_idx, val in enumerate(header_row, 1):
                    sheet2_clean.cell(row=clean_row_cursor, column=col_idx, value=val)
                clean_row_cursor += 1
                clean_header_written = True

            if data_start_idx < len(grid):
                for r_idx in range(data_start_idx, len(grid)):
                    row_data = grid[r_idx]
                    if any(c is not None and str(c).strip() != "" for c in row_data):
                        clean_row = [fund_name] + row_data
                        for col_idx, val in enumerate(clean_row, 1):
                            sheet2_clean.cell(row=clean_row_cursor, column=col_idx, value=val)
                        clean_row_cursor += 1

        output_path = os.path.join(self.output_dir, output_filename)
        out_wb.save(output_path)
        return {"status": "success", "total_processed_files": total_processed, "clean_total_rows": clean_row_cursor - 1, "output_file_path": output_path}

    def _detect_table_bounds(self, grid: List[List[Any]]) -> Tuple[Optional[int], int]:
        if not grid:
            return None, 0
        header_keywords = ["نام شرکت", "کد نماد", "نماد", "تعداد", "بهای تمام شده", "ارزش بازار"]
        for i, row in enumerate(grid[:15]):
            row_str = " ".join([str(c) for c in row if c is not None])
            if any(kw in row_str for kw in header_keywords):
                return i, i + 1
        return 0, 1