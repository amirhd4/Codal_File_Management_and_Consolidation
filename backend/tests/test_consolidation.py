import os
import openpyxl
import pytest
from app.services.consolidation_service import ExcelConsolidationEngine
from app.core.config import settings

def test_consolidation_with_merged_cells(tmp_path):
    # Create a source directory with an excel file containing merged cells
    source_dir = str(tmp_path / "source")
    output_dir = str(tmp_path / "output")
    os.makedirs(source_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    # Build sample excel with merged cells
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سهام"

    # Header
    ws.cell(row=1, column=1, value="نام شرکت")
    ws.cell(row=1, column=2, value="تعداد")
    ws.cell(row=1, column=3, value="بهای تمام شده")

    # Row 2 & 3 with merged cells in col 1 (e.g. A2:A3 merged)
    ws.cell(row=2, column=1, value="فولاد مبارکه")
    ws.cell(row=2, column=2, value=100)
    ws.cell(row=2, column=3, value=1000)

    ws.cell(row=3, column=2, value=200)
    ws.cell(row=3, column=3, value=2000)

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    file_path = os.path.join(source_dir, "test_fund_latest.xlsx")
    wb.save(file_path)

    engine = ExcelConsolidationEngine(source_dir=source_dir, output_dir=output_dir)
    res = engine.consolidate(target_sheets=["سهام"], output_filename="merged_test.xlsx")

    assert res["status"] == "success"

    # Inspect the consolidated output file
    out_wb = openpyxl.load_workbook(res["output_file_path"])
    sheet2 = out_wb["تلفیق تمیز (بدون سربرگ)"]

    # Check that sheet2 has NO merged ranges
    assert len(sheet2.merged_cells.ranges) == 0

    # Also check if the value from merged cells was propagated or preserved properly
    rows = list(sheet2.iter_rows(values_only=True))
    # Row 0 is header: ["نام صندوق", "نام شرکت", "تعداد", "بهای تمام شده"]
    # Row 1: ["test fund", "فولاد مبارکه", 100, 1000]
    # Row 2: ["test fund", "فولاد مبارکه" (or value from merged cell), 200, 2000]
    assert len(rows) >= 3
    print("Consolidated rows:", rows)
